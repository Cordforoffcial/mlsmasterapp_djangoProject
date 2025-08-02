from django.shortcuts import render, redirect
from django.http.response import HttpResponse, JsonResponse
from .models import *
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from openpyxl import Workbook
from openpyxl.styles import *
import decimal
from django.views.decorators.csrf import csrf_exempt
from django.core.exceptions import ValidationError
import json
from datetime import datetime
from django.contrib.auth import authenticate, login
from django.db import transaction

@csrf_exempt
def advanced_materials_testing(request):
    if request.method == 'POST':
        try:
            # Parse request body
            try:
                data = json.loads(request.body) if request.body else request.POST
            except json.JSONDecodeError:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid JSON data format'
                }, status=400)
            
            # Validate data structure
            if not isinstance(data, dict):
                return JsonResponse({
                    'status': 'error',
                    'message': 'Request data must be a JSON object'
                }, status=400)
            
            # Get structured data from the form
            sample_data = data.get('sample_parameters', {})
            water_data = data.get('water_system', {})
            scale_data = data.get('scale_load_measurements', {})
            
            # Extract required values for calculations
            mass = sample_data.get('mass')
            length = sample_data.get('length')
            utn_scale = scale_data.get('utn_scale')
            yield_load_main_scale = scale_data.get('yield_load_main_scale')
            yield_load_counter_part = scale_data.get('yield_load_counter_part')
            tensile_load_main_scale = scale_data.get('tensile_load_main_scale')
            tensile_load_counter_part = scale_data.get('tensile_load_counter_part')
            
            # Validate required fields for calculations
            required_fields = {
                'mass': mass,
                'length': length,
                'utn_scale': utn_scale,
                'yield_load_main_scale': yield_load_main_scale,
                'yield_load_counter_part': yield_load_counter_part,
                'tensile_load_main_scale': tensile_load_main_scale,
                'tensile_load_counter_part': tensile_load_counter_part
            }
            
            missing_fields = [field for field, value in required_fields.items() if not value]
            if missing_fields:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Missing required fields for calculations: {", ".join(missing_fields)}'
                }, status=400)
            
            # Convert to float
            try:
                mass = float(mass)
                length = float(length)
                utn_scale = float(utn_scale)
                yield_load_main_scale = float(yield_load_main_scale)
                yield_load_counter_part = float(yield_load_counter_part)
                tensile_load_main_scale = float(tensile_load_main_scale)
                tensile_load_counter_part = float(tensile_load_counter_part)
            except (ValueError, TypeError) as e:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Invalid numeric values: {str(e)}'
                }, status=400)
            
            # Create a new SampleParameters instance if provided
            sample_parameters = None
            if all(sample_data.get(field) for field in ['sample_number', 'heat_number']):
                sample_parameters = SampleParameters.objects.create(
                    sample_number=sample_data['sample_number'],
                    heat_number=sample_data['heat_number'],
                    mass=mass,
                    length=length
                )

            # Calculate and store stress results
            from .stress_calculator import calculate_and_store_stress
            
            result = calculate_and_store_stress(
                mass=mass,
                length=length,
                utn_scale=utn_scale,
                yield_load_main_scale=yield_load_main_scale,
                yield_load_counter_part=yield_load_counter_part,
                tensile_load_main_scale=tensile_load_main_scale,
                tensile_load_counter_part=tensile_load_counter_part,
                sample_parameters=sample_parameters
            )
            
            if not result['success']:
                return JsonResponse({
                    'status': 'error',
                    'message': result['error']
                }, status=400)
            
            # Extract calculated values
            calculated_values = result['calculated_values']
            yield_stress = calculated_values['yield_stress']
            tensile_stress = calculated_values['tensile_stress']
            division_type = calculated_values['division_type']
            
            # Store related data if provided
            if water_data:
                WaterSystem.objects.create(
                    water_pressure_in=float(water_data.get('water_pressure_in', 0)),
                    water_pressure_out=float(water_data.get('water_pressure_out', 0)),
                    water_in_temperature=float(water_data.get('water_in_temperature', 0)),
                    water_out_temperature=float(water_data.get('water_out_temperature', 0))
                )
                
            if scale_data:
                ScaleLoadMeasurements.objects.create(
                    utn_scale=utn_scale,
                    yield_load_main_scale=yield_load_main_scale,
                    yield_load_counter_part=yield_load_counter_part,
                    tensile_load_main_scale=tensile_load_main_scale,
                    tensile_load_counter_part=tensile_load_counter_part
                )
            else:
                return JsonResponse({
                    'status': 'error',
                    'message': f'Invalid UTN scale value: {utn_scale}. Must be 10, 25, 50, or 100.'
                }, status=400)
            
            # Save data to database
            sample = None
            water_system = None
            scale_load = None
            
            # Create SampleParameters record
            if sample_data.get('mass') and sample_data.get('length'):
                sample_kwargs = {
                    'mass': mass,
                    'length': length
                }
                if sample_data.get('sample_number'):
                    sample_kwargs['sample_number'] = sample_data.get('sample_number')
                if sample_data.get('heat_number'):
                    sample_kwargs['heat_number'] = sample_data.get('heat_number')
                
                sample = SampleParameters.objects.create(**sample_kwargs)
            
            # Create WaterSystem record if water data exists
            if any(water_data.values()):
                water_kwargs = {}
                for field_name, field_value in water_data.items():
                    if field_value:
                        water_kwargs[field_name] = float(field_value)
                
                if water_kwargs:
                    water_system = WaterSystem.objects.create(**water_kwargs)
            
            # Create ScaleLoadMeasurements record
            scale_kwargs = {
                'utn_scale': utn_scale,
                'yield_load_main_scale': yield_load_main_scale,
                'yield_load_counter_part': yield_load_counter_part,
                'tensile_load_main_scale': tensile_load_main_scale,
                'tensile_load_counter_part': tensile_load_counter_part
            }
            scale_load = ScaleLoadMeasurements.objects.create(**scale_kwargs)
            
            # Store calculation results in session for analysis_results view
            request.session['calculation_results'] = {
                'mass_per_meter': round(calculated_values['mass_per_meter'], 4),
                'cross_section_area': round(calculated_values['cross_section_area'], 2),
                'yield_machine_reading': round(calculated_values['yield_machine_reading'], 1),
                'tensile_machine_reading': round(calculated_values['tensile_machine_reading'], 1),
                'yield_stress': round(yield_stress, 2),
                'tensile_stress': round(tensile_stress, 2),
                'division_type': division_type,
                'utn_scale': utn_scale,
                'sample_id': sample.id if sample else None,
                'water_system_id': water_system.id if water_system else None,
                'scale_load_id': scale_load.id if scale_load else None
            }
            
            return JsonResponse({
                'status': 'success',
                'message': 'Data saved and calculations completed successfully',
                'redirect_url': '/analysis_results/',
                'data': {
                    'sample_id': sample.id if sample else None,
                    'water_system_id': water_system.id if water_system else None,
                    'scale_load_id': scale_load.id if scale_load else None
                }
            })
            
        except Exception as e:
            import traceback
            print(f"Error in advanced_materials_testing view: {str(e)}")
            print(traceback.format_exc())
            return JsonResponse({
                'status': 'error',
                'message': f'Server error: {str(e)}'
            }, status=500)
    
    # For GET requests, just render the template
    return render(request, 'advanced_materials_testing/advanced_materials_testing.html')


def analysis_results(request):
    """Display calculation results"""
    calculation_results = request.session.get('calculation_results')
    
    if not calculation_results:
        # If no calculation results in session, redirect back to form
        return redirect('advanced_materials_testing')
    
    context = {
        'results': calculation_results
    }
    
    return render(request, 'analysis_results/analysis_results.html', context)



def is_valid_queryparam(param):
    return param != '' and param is not None


def countries_gdp_list(request):
    qs = CountryGDP.objects.order_by('name')

    name = request.GET.get('name')
    year = request.GET.get('year')

    request.session['name'] = name
    request.session['year'] = year

    if is_valid_queryparam(name):
        qs = qs.filter(name__icontains=name)

    if is_valid_queryparam(year):
        qs = qs.filter(year=year)

    page = request.GET.get('page', 1)
    paginator = Paginator(qs, 30)

    try:
        qs = paginator.page(page)
    except PageNotAnInteger:
        qs = paginator.page(1)
    except EmptyPage:
        qs = paginator.page(paginator.num_pages)

    context = {
        'countries_list': qs,
        'name': name,
        'year':year,
    }
    return render(request, "excelexport/countries_list.html", context)


@csrf_exempt
def home(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body) if request.body else request.POST
            
            # Extract data from the request
            date_str = data.get('date')
            batch_number = data.get('batchNumber')
            section = data.get('section')

            # Validate required fields
            if not all([date_str, batch_number, section]):
                return JsonResponse({
                    'status': 'error',
                    'message': 'All fields are required'
                }, status=400)

            try:
                # Parse the date string
                inspection_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return JsonResponse({
                    'status': 'error',
                    'message': 'Invalid date format'
                }, status=400)

            # Create new inspection report
            inspection = InspectionReport.objects.create(
                date=inspection_date,
                batch_number=batch_number,
                section=section
            )

            return JsonResponse({
                'status': 'success',
                'message': 'Inspection report created successfully',
                'data': {
                    'id': inspection.id,
                    'date': inspection.date.isoformat(),
                    'batch_number': inspection.batch_number,
                    'section': inspection.section,
                    'created_at': inspection.created_at.isoformat()
                }
            })

        except json.JSONDecodeError:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid JSON data'
            }, status=400)
        
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': str(e)
            }, status=500)

    # For GET requests, just render the form
    return render(request, 'home/home.html')


def countries_gdp_excel(request):
    qs = CountryGDP.objects.order_by('name')

    if 'name' in request.session:
        name = request.session['name']
    else:
        name = None

    if 'year' in request.session:
        year = request.session['year']
    else:
        year = None

    if is_valid_queryparam(name):
        qs = qs.filter(name__icontains=name)

    if is_valid_queryparam(year):
        qs = qs.filter(year=year)

    if year is None or year == '':
        year = "2013 - 2016"
    else:
        year = year

    if name is None or name == '':
        name = "All Countries"
    else:
        name = name

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
    response['Content-Disposition'] = 'attachment; filename="' + 'Countries GDP List' +'.xlsx"'
    workbook = Workbook()

    worksheet = workbook.active

    worksheet.merge_cells('A1:D1')
    worksheet.merge_cells('A2:D2')
    first_cell = worksheet['A1']
    first_cell.value = "Countries GDP List" + " " + year
    first_cell.fill = PatternFill("solid", fgColor="246ba1")
    first_cell.font  = Font(bold=True, color="F7F6FA")
    first_cell.alignment = Alignment(horizontal="center", vertical="center")

    second_cell = worksheet['A2']
    second_cell.value = name
    second_cell.font  = Font(bold=True, color="246ba1")
    second_cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.title = 'Countries GDP List' + " " + year

    # Define the titles for columns
    columns = ['Country Name','Country Code','Year', 'Value in USD']
    row_num = 3

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.fill = PatternFill("solid", fgColor="50C878")
        cell.font  = Font(bold=True, color="F7F6FA")
        third_cell = worksheet['D3']
        third_cell.alignment = Alignment(horizontal="right")

    for countries in qs:
        row_num += 1

        # Define the data for each cell in the row
        row = [countries.name,countries.code,countries.year,countries.value]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            if isinstance(cell_value, decimal.Decimal):
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    workbook.save(response)
    return response

def mechanical_inspection_report(request):
    """
    View function for the mechanical inspection report page.
    """
    # Fetch all stress calculations with related data
    stress_calculations = StressCalculation.objects.select_related(
        'sample_parameters', 'inspection_report'
    ).order_by('created_at')[:15]  # Get oldest 15 records in ascending time order
    
    # Get header data from the latest inspection report
    header_data = {
        'section': '',
        'lot_no': '',
        'date_of_rolling': ''
    }
    
    latest_inspection = InspectionReport.objects.order_by('-created_at').first()
    if latest_inspection:
        header_data['section'] = latest_inspection.section
        header_data['lot_no'] = latest_inspection.batch_number
        header_data['date_of_rolling'] = latest_inspection.date.strftime('%Y-%m-%d')
    
    # Prepare data for the template
    table_data = []
    for i, calc in enumerate(stress_calculations, 1):
        # Get sample parameters if available
        sample = calc.sample_parameters
        
        row_data = {
            'serial_no': i,
            'heat_number': sample.heat_number if sample else '',
            'time': calc.created_at.strftime('%H:%M:%S') if calc.created_at else '',
            'mass': f"{sample.mass:.2f}" if sample and sample.mass else '',
            'length': f"{sample.length:.3f}" if sample and sample.length else '',
            'mass_per_meter': f"{calc.mass_per_meter:.4f}" if calc.mass_per_meter else '',
            'cross_section_area': f"{calc.cross_section_area:.2f}" if calc.cross_section_area else '',
            'calibrated_yield_load': f"{calc.yield_machine_reading:.1f}" if calc.yield_machine_reading else '',
            'yield_stress': f"{calc.yield_stress:.2f}" if calc.yield_stress else '',
            'calibrated_tensile_load': f"{calc.tensile_machine_reading:.1f}" if calc.tensile_machine_reading else '',
            'tensile_stress': f"{calc.tensile_stress:.2f}" if calc.tensile_stress else '',
            'elongation': '',  # This field is not in the current model
            'bend_test': ''    # This field is not in the current model
        }
        table_data.append(row_data)
    
    # Calculate summary statistics if we have data
    summary_data = {}
    if table_data:
        # Extract numeric values for calculations
        numeric_fields = ['mass', 'length', 'mass_per_meter', 'cross_section_area', 
                         'calibrated_yield_load', 'yield_stress', 'calibrated_tensile_load', 'tensile_stress']
        
        for field in numeric_fields:
            values = []
            for row in table_data:
                try:
                    if row[field]:
                        values.append(float(row[field]))
                except (ValueError, TypeError):
                    continue
            
            if values:
                summary_data[field] = {
                    'min': f"{min(values):.2f}",
                    'max': f"{max(values):.2f}",
                    'avg': f"{sum(values)/len(values):.2f}"
                }
            else:
                summary_data[field] = {'min': '', 'max': '', 'avg': ''}
    
    context = {
        'table_data': table_data,
        'summary_data': summary_data,
        'total_records': len(table_data),
        'header_data': header_data
    }
    
    return render(request, 'mechanical_inspection/mechanical_properties_report.html', context)


@csrf_exempt
def reset_database(request):
    """
    Secure database reset view that requires admin authentication
    """
    if request.method != 'POST':
        return JsonResponse({
            'status': 'error',
            'message': 'Only POST method allowed'
        }, status=405)
    
    try:
        # Parse request data
        data = json.loads(request.body) if request.body else {}
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        # Validate input
        if not username:
            return JsonResponse({
                'status': 'error',
                'message': 'Username is required',
                'field': 'username'
            }, status=400)
        
        if not password:
            return JsonResponse({
                'status': 'error',
                'message': 'Password is required',
                'field': 'password'
            }, status=400)
        
        # Authenticate user
        user = authenticate(username=username, password=password)
        
        if user is None:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid username or password',
                'field': 'username'
            }, status=401)
        
        # Check if user is admin/staff
        if not user.is_staff and not user.is_superuser:
            return JsonResponse({
                'status': 'error',
                'message': 'Only admin users can reset the database',
                'field': 'username'
            }, status=403)
        
        # Perform database reset in a transaction
        try:
            with transaction.atomic():
                # Delete all data from all models
                CountryGDP.objects.all().delete()
                InspectionReport.objects.all().delete()
                SampleParameters.objects.all().delete()
                WaterSystem.objects.all().delete()
                ScaleLoadMeasurements.objects.all().delete()
                StressCalculation.objects.all().delete()
                
                # Log the reset action (optional)
                print(f"Database reset performed by user: {username} at {datetime.now()}")
                
            return JsonResponse({
                'status': 'success',
                'message': 'All Data Reset Successfully Completed'
            })
            
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'Database reset failed: {str(e)}'
            }, status=500)
            
    except json.JSONDecodeError:
        return JsonResponse({
            'status': 'error',
            'message': 'Invalid JSON data'
        }, status=400)
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Server error: {str(e)}'
        }, status=500)


@csrf_exempt
def user_login(request):
    """
    User login view for the popup login system
    """
    if request.method != 'POST':
        return JsonResponse({
            'status': 'error',
            'message': 'Only POST method allowed'
        }, status=405)
    
    try:
        # Parse request data
        data = json.loads(request.body) if request.body else {}
        username = data.get('username', '').strip()
        password = data.get('password', '')
        
        # Validate input
        if not username:
            return JsonResponse({
                'status': 'error',
                'message': 'Username is required',
                'field': 'username'
            }, status=400)
        
        if not password:
            return JsonResponse({
                'status': 'error',
                'message': 'Password is required',
                'field': 'password'
            }, status=400)
        
        # Authenticate user
        user = authenticate(username=username, password=password)
        
        if user is None:
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid username or password',
                'field': 'username'
            }, status=401)
        
        if not user.is_active:
            return JsonResponse({
                'status': 'error',
                'message': 'Account is disabled',
                'field': 'username'
            }, status=401)
        
        # Log the user in
        login(request, user)
        
        # Log the login action
        print(f"User login successful: {username} at {datetime.now()}")
        
        return JsonResponse({
            'status': 'success',
            'message': 'Successfully logged in',
            'username': user.username,
            'user_id': user.id,
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser
        })
            
    except json.JSONDecodeError:
        return JsonResponse({
            'status': 'error',
            'message': 'Invalid JSON data'
        }, status=400)
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Server error: {str(e)}'
        }, status=500)